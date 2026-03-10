"""
Mise HERo Finance - Core App Views
===================================
Views for user management and authentication.
"""

from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, generics, serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny, IsAdminUser, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenObtainPairView

from .models import AuditLog, User
from .permissions import IsAdminOrSelf
from .serializers import (AdminUserUpdateSerializer, AuditLogSerializer,
                          CustomTokenObtainPairSerializer,
                          PasswordChangeSerializer, UserCreateSerializer,
                          UserDetailSerializer, UserSerializer,
                          UserUpdateSerializer)

# =============================================================================
# AUTHENTICATION VIEWS
# =============================================================================


class CustomTokenObtainPairView(TokenObtainPairView):
    """
    Custom JWT login endpoint that returns user info along with tokens.

    POST /api/v1/auth/token/
    {
        "email": "user@example.com",
        "password": "password123"
    }

    Returns:
    {
        "access": "...",
        "refresh": "...",
        "user": {...}
    }
    """

    serializer_class = CustomTokenObtainPairSerializer


class LogoutView(APIView):
    """
    Logout endpoint that blacklists the refresh token.

    POST /api/v1/auth/logout/
    {
        "refresh": "..."
    }
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data.get("refresh")
            if refresh_token:
                token = RefreshToken(refresh_token)
                token.blacklist()

            return Response({"success": True, "message": "Úspěšně odhlášeno."})
        except Exception as e:
            return Response(
                {"success": False, "error": str(e)}, status=status.HTTP_400_BAD_REQUEST
            )


class RegisterView(generics.CreateAPIView):
    """
    User registration endpoint.

    POST /api/v1/auth/register/
    {
        "email": "newuser@example.com",
        "password": "password123",
        "password_confirm": "password123",
        "first_name": "Jan",
        "last_name": "Novák"
    }
    """

    serializer_class = UserCreateSerializer
    permission_classes = [
        AllowAny
    ]  # Change to IsAdminUser if registration should be admin-only

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()

        # Generate tokens for new user
        refresh = RefreshToken.for_user(user)

        return Response(
            {
                "success": True,
                "message": "Uživatel úspěšně vytvořen.",
                "user": UserSerializer(user).data,
                "tokens": {
                    "access": str(refresh.access_token),
                    "refresh": str(refresh),
                },
            },
            status=status.HTTP_201_CREATED,
        )


# =============================================================================
# USER VIEWS
# =============================================================================


class CurrentUserView(APIView):
    """
    Get or update current authenticated user.

    GET /api/v1/auth/me/
    PATCH /api/v1/auth/me/
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Get current user profile."""
        serializer = UserDetailSerializer(request.user)
        return Response(serializer.data)

    def patch(self, request):
        """Update current user profile."""
        serializer = UserUpdateSerializer(request.user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(UserDetailSerializer(request.user).data)


class PasswordChangeView(APIView):
    """
    Change password for current user.

    POST /api/v1/auth/change-password/
    {
        "current_password": "oldpassword",
        "new_password": "newpassword123",
        "new_password_confirm": "newpassword123"
    }
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = PasswordChangeSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response({"success": True, "message": "Heslo bylo úspěšně změněno."})


class ForgotPasswordView(APIView):
    """
    Public endpoint for password reset.
    Generates a new random password and sends it via email.

    POST /api/v1/auth/forgot-password/
    {"email": "user@example.com"}
    """

    permission_classes = [AllowAny]

    def post(self, request):
        import secrets
        import string

        from django.conf import settings
        from django.core.mail import send_mail

        email = request.data.get("email", "").strip().lower()

        # Always return success to prevent email enumeration
        success_msg = (
            "Pokud je tento email registrován v systému, "
            "bylo na něj odesláno nové heslo."
        )

        if not email:
            return Response(
                {"success": False, "error": "Zadejte emailovou adresu."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            user = User.objects.get(email=email, is_active=True)
        except User.DoesNotExist:
            # Don't reveal whether email exists
            return Response({"success": True, "message": success_msg})

        # Generate new password
        alphabet = string.ascii_letters + string.digits
        new_password = "".join(secrets.choice(alphabet) for _ in range(10))
        user.set_password(new_password)
        user.save()

        # Try to send email
        email_sent = False
        try:
            send_mail(
                subject="HeroWizzard - Nové heslo",
                message=(
                    f"Dobrý den,\n\n"
                    f"Vaše heslo bylo resetováno.\n\n"
                    f"Nové heslo: {new_password}\n\n"
                    f"Po přihlášení si heslo můžete změnit.\n\n"
                    f"HeroWizzard"
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[user.email],
                fail_silently=False,
            )
            email_sent = True
        except Exception:
            pass

        if email_sent:
            return Response({"success": True, "message": success_msg})
        else:
            return Response({
                "success": True,
                "message": (
                    "Email se nepodařilo odeslat (SMTP není nakonfigurován). "
                    "Kontaktujte administrátora pro reset hesla."
                ),
                "email_sent": False,
            })


class UserViewSet(viewsets.ModelViewSet):
    """
    API endpoint for user management (admin only).

    list: Get all users
    retrieve: Get single user
    create: Create new user
    update: Update user
    destroy: Deactivate user
    """

    queryset = User.objects.all().order_by("-created_at")
    permission_classes = [IsAuthenticated, IsAdminUser]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["role", "is_active", "primary_kmen"]
    search_fields = ["email", "first_name", "last_name"]
    ordering_fields = ["email", "created_at", "last_login"]
    ordering = ["-created_at"]

    def get_serializer_class(self):
        if self.action == "create":
            return UserCreateSerializer
        elif self.action in ["update", "partial_update"]:
            # Admin can change role/is_active; non-admin can only change profile
            if self.request.user.is_staff or self.request.user.role == "admin":
                return AdminUserUpdateSerializer
            return UserUpdateSerializer
        elif self.action == "retrieve":
            return UserDetailSerializer
        return UserSerializer

    def perform_update(self, serializer):
        """Prevent removing admin role from last admin."""
        instance = serializer.instance
        new_role = serializer.validated_data.get("role")
        if (
            instance.role == "admin"
            and new_role
            and new_role != "admin"
        ):
            active_admins = User.objects.filter(
                role="admin", is_active=True
            ).count()
            if active_admins <= 1:
                raise serializers.ValidationError(
                    "Nelze odebrat roli administrátora poslednímu administrátorovi."
                )
        serializer.save()

    def get_permissions(self):
        """Allow users to view/update their own profile."""
        if self.action in ["retrieve", "update", "partial_update"]:
            return [IsAuthenticated(), IsAdminOrSelf()]
        return super().get_permissions()

    def perform_destroy(self, instance):
        """Soft delete by deactivating user. Prevent deactivating last admin."""
        if instance.role == "admin":
            active_admins = User.objects.filter(
                role="admin", is_active=True
            ).count()
            if active_admins <= 1:
                raise serializers.ValidationError(
                    "Nelze deaktivovat posledního administrátora."
                )
        instance.is_active = False
        instance.save()

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        """
        Export users as Excel (.xlsx). Admin only.

        GET /api/v1/users/export-excel/
        """
        from datetime import date as date_cls
        from io import BytesIO

        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        qs = self.get_queryset()

        wb = Workbook()
        ws = wb.active
        ws.title = "Uživatelé"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        headers = ["Email", "Jméno", "Příjmení", "Role", "Stav", "Vytvořen", "Poslední přihlášení"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row_idx, u in enumerate(qs.iterator(), 2):
            ws.cell(row=row_idx, column=1, value=u.email)
            ws.cell(row=row_idx, column=2, value=u.first_name or "")
            ws.cell(row=row_idx, column=3, value=u.last_name or "")
            ws.cell(row=row_idx, column=4, value=u.get_role_display() if hasattr(u, 'get_role_display') else u.role)
            ws.cell(row=row_idx, column=5, value="Aktivní" if u.is_active else "Neaktivní")
            ws.cell(row=row_idx, column=6, value=u.created_at.strftime("%d.%m.%Y %H:%M") if hasattr(u, 'created_at') and u.created_at else "")
            ws.cell(row=row_idx, column=7, value=u.last_login.strftime("%d.%m.%Y %H:%M") if u.last_login else "")

        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(headers[col_idx - 1]))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

        ws.freeze_panes = "A2"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        today = date_cls.today().strftime("%d_%m_%Y")
        filename = f"HeroWizzardUzivateleARole{today}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=["post"])
    def activate(self, request, pk=None):
        """Reactivate a deactivated user."""
        user = self.get_object()
        user.is_active = True
        user.save()

        return Response(
            {"success": True, "message": f"Uživatel {user.email} byl aktivován."}
        )

    @action(detail=True, methods=["post"])
    def reset_password(self, request, pk=None):
        """
        Admin endpoint to reset user password.
        Generates a random password and returns it.
        """
        import secrets
        import string

        user = self.get_object()

        # Generate random password
        alphabet = string.ascii_letters + string.digits
        new_password = "".join(secrets.choice(alphabet) for _ in range(12))

        user.set_password(new_password)
        user.save()

        return Response(
            {
                "success": True,
                "message": f"Heslo pro {user.email} bylo resetováno.",
                "temporary_password": new_password,
            }
        )

    @action(detail=True, methods=["post"])
    def set_password(self, request, pk=None):
        """
        Admin endpoint to set a specific password for a user.
        """
        user = self.get_object()
        new_password = request.data.get("new_password", "").strip()

        if not new_password:
            return Response(
                {"success": False, "error": "Heslo nesmí být prázdné."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if len(new_password) < 4:
            return Response(
                {"success": False, "error": "Heslo musí mít alespoň 4 znaky."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user.set_password(new_password)
        user.save()

        return Response(
            {
                "success": True,
                "message": f"Heslo pro {user.email} bylo nastaveno.",
            }
        )


# =============================================================================
# AUDIT LOG VIEWS
# =============================================================================


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for viewing audit logs (admin only).
    """

    queryset = AuditLog.objects.select_related("user").order_by("-timestamp")
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["action", "model_name", "user"]
    ordering_fields = ["timestamp"]
    ordering = ["-timestamp"]

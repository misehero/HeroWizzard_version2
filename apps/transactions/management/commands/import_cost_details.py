"""
Import CostDetail records from Excel files.

Usage:
    python manage.py import_cost_details [--clear]

Expects two Excel files in docs/test-data/:
    - docs/test-data/prijmy_druh_detail_poznamka.xlsx  (Příjmy — columns: V/N, Druh, Detail, Poznámky)
    - docs/test-data/vydaje_druh_detail_poznamka.xlsx  (Výdaje — columns: Druh, Detail, Poznámky)

Options:
    --clear   Delete all existing CostDetail records before importing
"""

import os
import re

from django.core.management.base import BaseCommand

from apps.transactions.models import CostDetail


def slugify_czech(text):
    """Create a slug from Czech text for use as primary key."""
    text = text.lower().strip()
    replacements = {
        "á": "a", "č": "c", "ď": "d", "é": "e", "ě": "e",
        "í": "i", "ň": "n", "ó": "o", "ř": "r", "š": "s",
        "ť": "t", "ú": "u", "ů": "u", "ý": "y", "ž": "z",
    }
    for cz, en in replacements.items():
        text = text.replace(cz, en)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")[:90]


class Command(BaseCommand):
    help = "Import CostDetail records from Excel files in docs/test-data/"

    def add_arguments(self, parser):
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Delete all existing CostDetail records before importing",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write("openpyxl is required: pip install openpyxl")
            return

        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        )))

        prijmy_file = os.path.join(base_dir, "docs", "test-data", "prijmy_druh_detail_poznamka.xlsx")
        vydaje_file = os.path.join(base_dir, "docs", "test-data", "vydaje_druh_detail_poznamka.xlsx")

        if not os.path.exists(prijmy_file):
            self.stderr.write(f"File not found: {prijmy_file}")
            return
        if not os.path.exists(vydaje_file):
            self.stderr.write(f"File not found: {vydaje_file}")
            return

        if options["clear"]:
            count = CostDetail.objects.count()
            CostDetail.objects.all().delete()
            self.stdout.write(f"Deleted {count} existing CostDetail records.")

        records = []

        # --- Příjmy ---
        wb = openpyxl.load_workbook(prijmy_file)
        ws = wb.active
        order = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            druh = str(row[1]).strip() if row[1] else ""
            detail = str(row[2]).strip() if row[2] else ""
            poznamka = str(row[3]).strip() if row[3] else ""
            if not druh:
                continue
            order += 1
            pk = f"prijmy-{slugify_czech(druh)}-{slugify_czech(detail)}"
            records.append(CostDetail(
                id=pk,
                druh_type="prijmy",
                druh_value=druh,
                detail=detail,
                poznamka=poznamka,
                sort_order=order,
            ))

        # --- Výdaje ---
        wb2 = openpyxl.load_workbook(vydaje_file)
        ws2 = wb2.active
        order = 0
        for row in ws2.iter_rows(min_row=2, values_only=True):
            druh = str(row[0]).strip() if row[0] else ""
            detail = str(row[1]).strip() if row[1] else ""
            poznamka = str(row[2]).strip() if row[2] else ""
            if not druh:
                continue
            order += 1
            pk = f"vydaje-{slugify_czech(druh)}-{slugify_czech(detail)}"
            records.append(CostDetail(
                id=pk,
                druh_type="vydaje",
                druh_value=druh,
                detail=detail,
                poznamka=poznamka,
                sort_order=order,
            ))

        # Bulk create, skip duplicates
        created = CostDetail.objects.bulk_create(records, ignore_conflicts=True)
        self.stdout.write(self.style.SUCCESS(
            f"Imported {len(records)} CostDetail records "
            f"({sum(1 for r in records if r.druh_type == 'prijmy')} příjmy, "
            f"{sum(1 for r in records if r.druh_type == 'vydaje')} výdaje)."
        ))

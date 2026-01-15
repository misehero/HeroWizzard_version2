"""
Management command to export transactions to CSV or Excel.

Usage:
    python manage.py export_transactions output.csv
    python manage.py export_transactions output.xlsx --format=excel
    python manage.py export_transactions output.csv --date-from=2024-01-01 --date-to=2024-12-31
    python manage.py export_transactions output.csv --status=zpracovano
"""

import csv

from django.core.management.base import BaseCommand, CommandError

from apps.transactions.models import Transaction


class Command(BaseCommand):
    help = "Export transactions to CSV or Excel file"

    def add_arguments(self, parser):
        parser.add_argument(
            "output_file",
            type=str,
            help="Path to output file",
        )
        parser.add_argument(
            "--format",
            type=str,
            choices=["csv", "excel"],
            default="csv",
            help="Output format (csv or excel)",
        )
        parser.add_argument(
            "--date-from",
            type=str,
            help="Filter from date (YYYY-MM-DD)",
        )
        parser.add_argument(
            "--date-to",
            type=str,
            help="Filter to date (YYYY-MM-DD)",
        )
        parser.add_argument(
            "--status",
            type=str,
            help="Filter by status",
        )
        parser.add_argument(
            "--projekt",
            type=str,
            help="Filter by project ID",
        )
        parser.add_argument(
            "--uncategorized-only",
            action="store_true",
            help="Export only uncategorized transactions",
        )

    def handle(self, *args, **options):
        output_path = options["output_file"]
        output_format = options["format"]

        # Build queryset
        qs = Transaction.objects.select_related(
            "projekt", "produkt", "podskupina"
        ).order_by("datum", "created_at")

        # Apply filters
        if options["date_from"]:
            qs = qs.filter(datum__gte=options["date_from"])
        if options["date_to"]:
            qs = qs.filter(datum__lte=options["date_to"])
        if options["status"]:
            qs = qs.filter(status=options["status"])
        if options["projekt"]:
            qs = qs.filter(projekt_id=options["projekt"])
        if options["uncategorized_only"]:
            from django.db.models import Q

            qs = qs.filter(Q(prijem_vydaj="") | Q(druh=""))

        count = qs.count()
        self.stdout.write(f"Exporting {count} transactions...")

        if count == 0:
            self.stdout.write(self.style.WARNING("No transactions to export"))
            return

        # Define columns
        columns = [
            ("Datum", lambda t: t.datum.strftime("%d.%m.%Y") if t.datum else ""),
            ("Účet", lambda t: t.ucet or ""),
            ("Typ", lambda t: t.typ or ""),
            ("Poznámka/Zpráva", lambda t: t.poznamka_zprava or ""),
            ("VS", lambda t: t.variabilni_symbol or ""),
            ("Částka", lambda t: str(t.castka).replace(".", ",")),
            ("Status", lambda t: t.get_status_display()),
            ("P/V", lambda t: t.prijem_vydaj),
            ("V/N", lambda t: t.vlastni_nevlastni),
            ("Daně", lambda t: "Ano" if t.dane else "Ne"),
            ("Druh", lambda t: t.druh or ""),
            ("Detail", lambda t: t.detail or ""),
            ("KMEN", lambda t: t.kmen or ""),
            ("MH%", lambda t: str(t.mh_pct).replace(".", ",")),
            ("ŠK%", lambda t: str(t.sk_pct).replace(".", ",")),
            ("XP%", lambda t: str(t.xp_pct).replace(".", ",")),
            ("FR%", lambda t: str(t.fr_pct).replace(".", ",")),
            ("Projekt", lambda t: t.projekt.name if t.projekt else ""),
            ("Produkt", lambda t: t.produkt.name if t.produkt else ""),
            ("Podskupina", lambda t: t.podskupina.name if t.podskupina else ""),
            ("Číslo protiúčtu", lambda t: t.cislo_protiuctu or ""),
            ("Název protiúčtu", lambda t: t.nazev_protiuctu or ""),
            ("Merchant", lambda t: t.nazev_merchanta or ""),
            ("ID transakce", lambda t: t.id_transakce or ""),
        ]

        if output_format == "csv":
            self._export_csv(output_path, qs, columns)
        else:
            self._export_excel(output_path, qs, columns)

        self.stdout.write(
            self.style.SUCCESS(f"Exported {count} transactions to {output_path}")
        )

    def _export_csv(self, output_path, queryset, columns):
        """Export to CSV file."""
        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")

            # Header
            writer.writerow([col[0] for col in columns])

            # Data
            for txn in queryset.iterator():
                row = [col[1](txn) for col in columns]
                writer.writerow(row)

    def _export_excel(self, output_path, queryset, columns):
        """Export to Excel file."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise CommandError(
                "openpyxl is required for Excel export. Install with: pip install openpyxl"
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transakce"

        # Header styling
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

        # Header row
        for col_idx, (name, _) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = header_font
            cell.fill = header_fill

        # Data rows
        for row_idx, txn in enumerate(queryset.iterator(), start=2):
            for col_idx, (_, getter) in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=getter(txn))

        # Adjust column widths
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(output_path)

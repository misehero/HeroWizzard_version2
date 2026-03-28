"""
Mise HERo Finance - Transactions ViewSets
==========================================
DRF ViewSets for all transaction-related API endpoints.
"""

from decimal import Decimal

from django.db.models import Count, F, Q, Sum
from django.db.models.functions import Coalesce, TruncMonth
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .filters import TransactionFilter
from .models import (CategoryRule, CostDetail, ImportBatch, Product,
                     ProductSubgroup, Project, Transaction,
                     TransactionAuditLog)
from .serializers import (CategoryRuleSerializer, CostDetailSerializer,
                          CSVUploadSerializer, ImportBatchSerializer,
                          ManualTransactionSerializer, MonthlyTrendSerializer,
                          ProductSerializer, ProductSubgroupDetailSerializer,
                          ProjectSerializer, TransactionAuditLogSerializer,
                          TransactionBulkUpdateSerializer,
                          TransactionDetailSerializer,
                          TransactionListSerializer,
                          TransactionStatsSerializer)
from .services import IDokladImporter, TransactionImporter

# =============================================================================
# LOOKUP VIEWSETS
# =============================================================================


class ProjectViewSet(viewsets.ModelViewSet):
    """
    API endpoint for Project lookup management.

    list: Get all projects (no pagination — lookup tables are small)
    retrieve: Get single project by ID
    create: Create new project (admin only)
    update: Update project (admin only)
    delete: Soft-delete project (admin only)
    """

    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    pagination_class = None
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name", "description"]
    ordering_fields = ["sort_order", "name", "created_at"]
    ordering = ["sort_order", "name"]

    def get_queryset(self):
        """Filter to active projects unless ?include_inactive=true or detail view."""
        qs = super().get_queryset()
        # Don't filter on detail actions (retrieve, update, partial_update, destroy)
        if self.action == "list" and not self.request.query_params.get(
            "include_inactive"
        ):
            qs = qs.filter(is_active=True)
        return qs

    def perform_destroy(self, instance):
        """Soft delete by setting is_active=False."""
        instance.is_active = False
        instance.save()

    @action(detail=True, methods=["post"], url_path="reorder")
    def reorder(self, request, pk=None):
        """Move item up or down. POST with {"direction": "up"} or {"direction": "down"}."""
        direction = request.data.get("direction")
        if direction not in ("up", "down"):
            return Response(
                {"error": "direction must be 'up' or 'down'"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        item = self.get_object()
        Model = item.__class__
        ordering = list(Model._meta.ordering or ["sort_order", "name"])
        items = list(Model.objects.order_by(*ordering))
        idx = next((i for i, x in enumerate(items) if x.pk == item.pk), None)
        if idx is None:
            return Response(
                {"error": "item not found"}, status=status.HTTP_404_NOT_FOUND
            )
        if direction == "up" and idx > 0:
            swap = items[idx - 1]
        elif direction == "down" and idx < len(items) - 1:
            swap = items[idx + 1]
        else:
            return Response({"status": "already at boundary"})
        # Swap sort_order values
        item.sort_order, swap.sort_order = swap.sort_order, item.sort_order
        # If sort_orders were equal, force differentiation
        if item.sort_order == swap.sort_order:
            if direction == "up":
                item.sort_order -= 1
            else:
                item.sort_order += 1
        item.save(update_fields=["sort_order", "updated_at"])
        swap.save(update_fields=["sort_order", "updated_at"])
        return Response({"status": "ok"})


class LookupsExcelExportView(APIView):
    """
    Export all lookups (projects, products, subgroups, cost details) as Excel (.xlsx) with multiple sheets.

    GET /api/v1/lookups/export-excel/
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        from datetime import date as date_cls
        from io import BytesIO

        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="2563EB", end_color="2563EB", fill_type="solid"
        )
        header_align = Alignment(horizontal="center", vertical="center")

        def style_headers(ws, headers):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

        def auto_width(ws, headers):
            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                max_length = len(str(headers[col_idx - 1]))
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
            ws.freeze_panes = "A2"

        wb = Workbook()

        # Sheet 1: Projects
        ws_proj = wb.active
        ws_proj.title = "Projekty"
        proj_headers = ["Název", "Popis", "Pořadí", "Aktivní", "Vytvořeno", "Změněno"]
        style_headers(ws_proj, proj_headers)
        for row_idx, p in enumerate(Project.objects.order_by("sort_order", "name"), 2):
            ws_proj.cell(row=row_idx, column=1, value=p.name)
            ws_proj.cell(row=row_idx, column=2, value=p.description or "")
            ws_proj.cell(row=row_idx, column=3, value=p.sort_order)
            ws_proj.cell(row=row_idx, column=4, value="Ano" if p.is_active else "Ne")
            ws_proj.cell(
                row=row_idx,
                column=5,
                value=p.created_at.strftime("%d.%m.%Y %H:%M") if p.created_at else "",
            )
            ws_proj.cell(
                row=row_idx,
                column=6,
                value=p.updated_at.strftime("%d.%m.%Y %H:%M") if p.updated_at else "",
            )
        auto_width(ws_proj, proj_headers)

        # Sheet 2: Products
        ws_prod = wb.create_sheet("Produkty")
        prod_headers = [
            "Název",
            "Popis",
            "Kategorie",
            "Pořadí",
            "Aktivní",
            "Vytvořeno",
            "Změněno",
        ]
        style_headers(ws_prod, prod_headers)
        for row_idx, p in enumerate(
            Product.objects.order_by("sort_order", "category", "name"), 2
        ):
            ws_prod.cell(row=row_idx, column=1, value=p.name)
            ws_prod.cell(row=row_idx, column=2, value=p.description or "")
            ws_prod.cell(row=row_idx, column=3, value=p.category or "")
            ws_prod.cell(row=row_idx, column=4, value=p.sort_order)
            ws_prod.cell(row=row_idx, column=5, value="Ano" if p.is_active else "Ne")
            ws_prod.cell(
                row=row_idx,
                column=6,
                value=p.created_at.strftime("%d.%m.%Y %H:%M") if p.created_at else "",
            )
            ws_prod.cell(
                row=row_idx,
                column=7,
                value=p.updated_at.strftime("%d.%m.%Y %H:%M") if p.updated_at else "",
            )
        auto_width(ws_prod, prod_headers)

        # Sheet 3: Subgroups
        ws_sub = wb.create_sheet("Podskupiny")
        sub_headers = [
            "Název",
            "Popis",
            "Produkt",
            "Pořadí",
            "Aktivní",
            "Vytvořeno",
            "Změněno",
        ]
        style_headers(ws_sub, sub_headers)
        for row_idx, s in enumerate(
            ProductSubgroup.objects.select_related("product").order_by(
                "product__name", "sort_order", "name"
            ),
            2,
        ):
            ws_sub.cell(row=row_idx, column=1, value=s.name)
            ws_sub.cell(row=row_idx, column=2, value=s.description or "")
            ws_sub.cell(
                row=row_idx, column=3, value=s.product.name if s.product else ""
            )
            ws_sub.cell(row=row_idx, column=4, value=s.sort_order)
            ws_sub.cell(row=row_idx, column=5, value="Ano" if s.is_active else "Ne")
            ws_sub.cell(
                row=row_idx,
                column=6,
                value=s.created_at.strftime("%d.%m.%Y %H:%M") if s.created_at else "",
            )
            ws_sub.cell(
                row=row_idx,
                column=7,
                value=s.updated_at.strftime("%d.%m.%Y %H:%M") if s.updated_at else "",
            )
        auto_width(ws_sub, sub_headers)

        # Sheet 4: Cost Details
        ws_cost = wb.create_sheet("Druhy nákladů")
        cost_headers = ["Typ", "Druh", "Detail", "Poznámka", "Pořadí", "Aktivní"]
        style_headers(ws_cost, cost_headers)
        for row_idx, c in enumerate(
            CostDetail.objects.order_by(
                "druh_type", "sort_order", "druh_value", "detail"
            ),
            2,
        ):
            ws_cost.cell(row=row_idx, column=1, value=c.get_druh_type_display())
            ws_cost.cell(row=row_idx, column=2, value=c.druh_value)
            ws_cost.cell(row=row_idx, column=3, value=c.detail or "")
            ws_cost.cell(row=row_idx, column=4, value=c.poznamka or "")
            ws_cost.cell(row=row_idx, column=5, value=c.sort_order)
            ws_cost.cell(row=row_idx, column=6, value="Ano" if c.is_active else "Ne")
        auto_width(ws_cost, cost_headers)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        today = date_cls.today().strftime("%d_%m_%Y")
        filename = f"HeroWizzardCiselniky{today}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class ProductViewSet(viewsets.ModelViewSet):
    """
    API endpoint for Product lookup management.
    Includes nested subgroups in response.
    """

    queryset = Product.objects.prefetch_related("subgroups")
    serializer_class = ProductSerializer
    pagination_class = None
    permission_classes = [IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["category", "is_active"]
    search_fields = ["name", "description"]
    ordering_fields = ["sort_order", "category", "name", "created_at"]
    ordering = ["sort_order", "category", "name"]

    def get_queryset(self):
        qs = super().get_queryset()
        if self.action == "list" and not self.request.query_params.get(
            "include_inactive"
        ):
            qs = qs.filter(is_active=True)
        return qs

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()

    @action(detail=True, methods=["post"], url_path="reorder")
    def reorder(self, request, pk=None):
        """Move item up or down."""
        direction = request.data.get("direction")
        if direction not in ("up", "down"):
            return Response(
                {"error": "direction must be 'up' or 'down'"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        item = self.get_object()
        Model = item.__class__
        ordering = list(Model._meta.ordering or ["sort_order", "name"])
        items = list(Model.objects.order_by(*ordering))
        idx = next((i for i, x in enumerate(items) if x.pk == item.pk), None)
        if idx is None:
            return Response(
                {"error": "item not found"}, status=status.HTTP_404_NOT_FOUND
            )
        if direction == "up" and idx > 0:
            swap = items[idx - 1]
        elif direction == "down" and idx < len(items) - 1:
            swap = items[idx + 1]
        else:
            return Response({"status": "already at boundary"})
        item.sort_order, swap.sort_order = swap.sort_order, item.sort_order
        if item.sort_order == swap.sort_order:
            if direction == "up":
                item.sort_order -= 1
            else:
                item.sort_order += 1
        item.save(update_fields=["sort_order", "updated_at"])
        swap.save(update_fields=["sort_order", "updated_at"])
        return Response({"status": "ok"})


class ProductSubgroupViewSet(viewsets.ModelViewSet):
    """
    API endpoint for ProductSubgroup lookup management.
    """

    queryset = ProductSubgroup.objects.select_related("product")
    serializer_class = ProductSubgroupDetailSerializer
    pagination_class = None
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["product", "is_active"]
    search_fields = ["name", "description"]

    def get_queryset(self):
        qs = super().get_queryset()
        if self.action == "list" and not self.request.query_params.get(
            "include_inactive"
        ):
            qs = qs.filter(is_active=True)
        return qs

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()

    @action(detail=True, methods=["post"], url_path="reorder")
    def reorder(self, request, pk=None):
        """Move item up or down within its product group."""
        direction = request.data.get("direction")
        if direction not in ("up", "down"):
            return Response(
                {"error": "direction must be 'up' or 'down'"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        item = self.get_object()
        # Subgroups reorder within their product
        items = list(
            ProductSubgroup.objects.filter(product=item.product).order_by(
                "sort_order", "name"
            )
        )
        idx = next((i for i, x in enumerate(items) if x.pk == item.pk), None)
        if idx is None:
            return Response(
                {"error": "item not found"}, status=status.HTTP_404_NOT_FOUND
            )
        if direction == "up" and idx > 0:
            swap = items[idx - 1]
        elif direction == "down" and idx < len(items) - 1:
            swap = items[idx + 1]
        else:
            return Response({"status": "already at boundary"})
        item.sort_order, swap.sort_order = swap.sort_order, item.sort_order
        if item.sort_order == swap.sort_order:
            if direction == "up":
                item.sort_order -= 1
            else:
                item.sort_order += 1
        item.save(update_fields=["sort_order", "updated_at"])
        swap.save(update_fields=["sort_order", "updated_at"])
        return Response({"status": "ok"})


class CostDetailViewSet(viewsets.ModelViewSet):
    """
    API endpoint for CostDetail (Druh/Detail) lookup management.
    """

    queryset = CostDetail.objects.all()
    serializer_class = CostDetailSerializer
    pagination_class = None
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["druh_type", "druh_value", "is_active"]
    search_fields = ["druh_value", "detail", "poznamka"]
    ordering = ["druh_type", "sort_order", "druh_value", "detail"]

    def get_queryset(self):
        qs = super().get_queryset()
        if self.request.query_params.get("include_inactive") != "true":
            if "is_active" not in self.request.query_params:
                qs = qs.filter(is_active=True)
        return qs

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()

    @action(detail=True, methods=["post"], url_path="reorder")
    def reorder(self, request, pk=None):
        """Move item up or down."""
        direction = request.data.get("direction")
        if direction not in ("up", "down"):
            return Response({"error": "direction must be 'up' or 'down'"}, status=400)
        item = self.get_object()
        items = list(
            CostDetail.objects.order_by("druh_type", "sort_order", "druh_value")
        )
        idx = next((i for i, x in enumerate(items) if x.pk == item.pk), None)
        if idx is None:
            return Response({"status": "not found"}, status=404)
        if direction == "up" and idx > 0:
            swap = items[idx - 1]
        elif direction == "down" and idx < len(items) - 1:
            swap = items[idx + 1]
        else:
            return Response({"status": "already at boundary"})
        item.sort_order, swap.sort_order = swap.sort_order, item.sort_order
        if item.sort_order == swap.sort_order:
            if direction == "up":
                item.sort_order -= 1
            else:
                item.sort_order += 1
        item.save(update_fields=["sort_order"])
        swap.save(update_fields=["sort_order"])
        return Response({"status": "ok"})


# =============================================================================
# TRANSACTION VIEWSET
# =============================================================================


class TransactionViewSet(viewsets.ModelViewSet):
    """
    API endpoint for Transaction management.

    list: Get paginated transactions with filters
    retrieve: Get single transaction with full details
    update: Update app columns (bank columns are read-only)
    bulk_update: Update multiple transactions at once
    stats: Get aggregated statistics
    trends: Get monthly trend data
    """

    queryset = Transaction.objects.select_related(
        "projekt", "produkt", "podskupina", "updated_by"
    ).order_by("-datum", "-created_at")
    permission_classes = [IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = TransactionFilter
    search_fields = [
        "poznamka_zprava",
        "nazev_protiuctu",
        "nazev_merchanta",
        "variabilni_symbol",
        "vlastni_poznamka",
    ]
    ordering_fields = [
        "datum",
        "castka",
        "status",
        "prijem_vydaj",
        "druh",
        "created_at",
    ]
    ordering = ["-datum", "-created_at"]

    def get_serializer_class(self):
        """Use list serializer for list action, detail for others."""
        if self.action == "list":
            return TransactionListSerializer
        return TransactionDetailSerializer

    def get_queryset(self):
        """Apply date range, is_active, and is_deleted filters."""
        qs = super().get_queryset()

        # Always exclude soft-deleted transactions
        qs = qs.filter(is_deleted=False)

        # Date range filtering
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")

        if date_from:
            qs = qs.filter(datum__gte=date_from)
        if date_to:
            qs = qs.filter(datum__lte=date_to)

        # Active/inactive filtering: only on list/stats/trends, not on retrieve/update
        if self.action in ("list", "stats", "trends"):
            show_inactive = self.request.query_params.get("show_inactive")
            if show_inactive not in ("true", "1"):
                qs = qs.filter(is_active=True)

        return qs

    @staticmethod
    def _format_audit_value(val):
        """Format a value for audit log display."""
        if val is None or val == "":
            return "—"
        if isinstance(val, bool):
            return "Ano" if val else "Ne"
        return str(val)

    def perform_update(self, serializer):
        """Set updated_by and create audit log on every update.

        Status rules:
        - Only admin/manager can explicitly change status
        - Any save by accountant/viewer forces status to 'ceka_na_schvaleni'
        """
        instance = serializer.instance
        user = self.request.user
        user_role = getattr(user, "role", "viewer")

        # Block status changes from non-admin/manager users
        if "status" in serializer.validated_data and user_role not in (
            "admin",
            "manager",
        ):
            from rest_framework.exceptions import PermissionDenied

            raise PermissionDenied("Pouze admin nebo manažer může měnit status.")

        # Capture changes before save
        fmt = self._format_audit_value
        changes = []
        for field_name, new_val in serializer.validated_data.items():
            old_val = getattr(instance, field_name)
            old_str = fmt(old_val)
            new_str = fmt(new_val)
            if old_str != new_str:
                try:
                    verbose = instance._meta.get_field(field_name).verbose_name
                except Exception:
                    verbose = field_name
                changes.append(f"{verbose}: {old_str} → {new_str}")

        # Auto-set status for accountant/viewer on any save
        extra_kwargs = {"updated_by": user}
        if user_role in ("accountant", "viewer"):
            extra_kwargs["status"] = Transaction.Status.CEKA_NA_SCHVALENI
            old_status = fmt(instance.status)
            new_status = fmt(Transaction.Status.CEKA_NA_SCHVALENI)
            if old_status != new_status:
                changes.append(f"Status: {old_status} → {new_status} (auto)")

        serializer.save(**extra_kwargs)

        if changes:
            TransactionAuditLog.objects.create(
                transaction=instance,
                user=user,
                action="Úprava",
                details="; ".join(changes),
            )

    @action(detail=True, methods=["get"], url_path="audit-log")
    def audit_log(self, request, pk=None):
        """
        Get audit log entries for a specific transaction.

        GET /api/v1/transactions/{id}/audit-log/
        """
        transaction = self.get_object()
        logs = TransactionAuditLog.objects.filter(
            transaction=transaction
        ).select_related("user")
        serializer = TransactionAuditLogSerializer(logs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["post"])
    def bulk_update(self, request):
        """
        Bulk update multiple transactions.

        POST /api/v1/transactions/bulk_update/
        {
            "ids": ["uuid1", "uuid2", ...],
            "status": "zpracovano",
            "projekt": "4cfuture"
        }
        """
        serializer = TransactionBulkUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        count = serializer.update(serializer.validated_data)

        return Response(
            {
                "success": True,
                "updated_count": count,
                "message": f"Úspěšně aktualizováno {count} transakcí.",
            }
        )

    @action(detail=False, methods=["get"])
    def stats(self, request):
        """
        Get aggregated transaction statistics.

        GET /api/v1/transactions/stats/?date_from=2024-01-01&date_to=2024-12-31
        """
        qs = self.get_queryset()

        # Basic aggregates
        totals = qs.aggregate(
            total_count=Count("id"),
            total_income=Coalesce(Sum("castka", filter=Q(castka__gt=0)), Decimal("0")),
            total_expense=Coalesce(Sum("castka", filter=Q(castka__lt=0)), Decimal("0")),
        )
        totals["net_balance"] = totals["total_income"] + totals["total_expense"]

        # By status
        status_counts = dict(
            qs.values("status")
            .annotate(count=Count("id"))
            .values_list("status", "count")
        )

        # By KMEN (weighted by percentage)
        by_kmen = {
            "MH": qs.aggregate(
                total=Coalesce(Sum(F("castka") * F("mh_pct") / 100), Decimal("0"))
            )["total"],
            "SK": qs.aggregate(
                total=Coalesce(Sum(F("castka") * F("sk_pct") / 100), Decimal("0"))
            )["total"],
            "XP": qs.aggregate(
                total=Coalesce(Sum(F("castka") * F("xp_pct") / 100), Decimal("0"))
            )["total"],
            "FR": qs.aggregate(
                total=Coalesce(Sum(F("castka") * F("fr_pct") / 100), Decimal("0"))
            )["total"],
        }

        # By Druh
        by_druh = dict(
            qs.exclude(druh="")
            .values("druh")
            .annotate(total=Sum("castka"))
            .values_list("druh", "total")
        )

        # Uncategorized
        uncategorized = qs.filter(Q(prijem_vydaj="") | Q(druh="")).aggregate(
            count=Count("id"), amount=Coalesce(Sum("castka"), Decimal("0"))
        )

        stats_data = {
            "total_count": totals["total_count"],
            "total_income": totals["total_income"],
            "total_expense": abs(totals["total_expense"]),
            "net_balance": totals["net_balance"],
            "by_status": status_counts,
            "by_kmen": by_kmen,
            "by_druh": by_druh,
            "uncategorized_count": uncategorized["count"],
            "uncategorized_amount": uncategorized["amount"],
        }

        serializer = TransactionStatsSerializer(stats_data)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def trends(self, request):
        """
        Get monthly trend data.

        GET /api/v1/transactions/trends/?months=12
        """
        months = int(request.query_params.get("months", 12))
        qs = self.get_queryset()

        # Get monthly aggregates
        monthly_data = (
            qs.annotate(month=TruncMonth("datum"))
            .values("month")
            .annotate(
                income=Coalesce(Sum("castka", filter=Q(castka__gt=0)), Decimal("0")),
                expense=Coalesce(Sum("castka", filter=Q(castka__lt=0)), Decimal("0")),
                transaction_count=Count("id"),
            )
            .order_by("-month")[:months]
        )

        trends = []
        for item in monthly_data:
            trends.append(
                {
                    "month": item["month"],
                    "income": item["income"],
                    "expense": abs(item["expense"]),
                    "net": item["income"] + item["expense"],
                    "transaction_count": item["transaction_count"],
                }
            )

        serializer = MonthlyTrendSerializer(trends, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def export(self, request):
        """
        Export transactions as CSV.

        GET /api/v1/transactions/export/?format=csv
        """
        import csv

        from django.http import HttpResponse

        qs = self.filter_queryset(self.get_queryset())

        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = (
            'attachment; filename="transactions_export.csv"'
        )

        writer = csv.writer(response, delimiter=";")

        # Header row
        headers = [
            "Datum",
            "Účet",
            "Typ",
            "Poznámka/Zpráva",
            "VS",
            "Částka",
            "Měna",
            "Zdroj",
            "Vyplaceno",
            "Status",
            "P/V",
            "V/N",
            "Daně",
            "Druh",
            "Detail",
            "KMEN",
            "MH%",
            "ŠK%",
            "XP%",
            "FR%",
            "Projekt",
            "Produkt",
            "Podskupina",
        ]
        writer.writerow(headers)

        # Data rows
        for t in qs.iterator():
            writer.writerow(
                [
                    t.datum.strftime("%d.%m.%Y") if t.datum else "",
                    t.ucet,
                    t.typ,
                    t.poznamka_zprava,
                    t.variabilni_symbol,
                    str(t.castka).replace(".", ","),
                    t.mena,
                    t.get_zdroj_transakce_display(),
                    "Ano" if t.vyplaceno else "Ne",
                    t.get_status_display(),
                    t.prijem_vydaj,
                    t.vlastni_nevlastni,
                    "Ano" if t.dane else "Ne",
                    t.druh,
                    t.detail,
                    t.kmen,
                    str(t.mh_pct).replace(".", ","),
                    str(t.sk_pct).replace(".", ","),
                    str(t.xp_pct).replace(".", ","),
                    str(t.fr_pct).replace(".", ","),
                    t.projekt.name if t.projekt else "",
                    t.produkt.name if t.produkt else "",
                    t.podskupina.name if t.podskupina else "",
                ]
            )

        return response

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        """
        Export filtered transactions as Excel (.xlsx).

        GET /api/v1/transactions/export-excel/?status=...&prijem_vydaj=...&date_from=...&date_to=...&search=...
        """
        from io import BytesIO

        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        qs = self.filter_queryset(self.get_queryset())
        # Always exclude inactive transactions from Excel export
        qs = qs.filter(is_active=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Transakce"

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="2563EB", end_color="2563EB", fill_type="solid"
        )
        header_align = Alignment(horizontal="center", vertical="center")

        headers = [
            "Datum",
            "Účet",
            "Typ",
            "Poznámka/Zpráva",
            "VS",
            "Částka",
            "Měna",
            "Zdroj",
            "Vyplaceno",
            "Číslo protiúčtu",
            "Název protiúčtu",
            "Název obchodníka",
            "Město",
            "Status",
            "P/V",
            "V/N",
            "Daně",
            "Druh",
            "Detail",
            "Zodpovědná osoba",
            "KMEN",
            "MH%",
            "ŠK%",
            "XP%",
            "FR%",
            "Projekt",
            "Produkt",
            "Podskupina",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Data rows
        for row_idx, t in enumerate(qs.iterator(), 2):
            ws.cell(
                row=row_idx,
                column=1,
                value=t.datum.strftime("%d.%m.%Y") if t.datum else "",
            )
            ws.cell(row=row_idx, column=2, value=t.ucet or "")
            ws.cell(row=row_idx, column=3, value=t.typ or "")
            ws.cell(row=row_idx, column=4, value=t.poznamka_zprava or "")
            ws.cell(row=row_idx, column=5, value=t.variabilni_symbol or "")
            ws.cell(row=row_idx, column=6, value=float(t.castka) if t.castka else 0)
            ws.cell(row=row_idx, column=7, value=t.mena or "CZK")
            ws.cell(row=row_idx, column=8, value=t.get_zdroj_transakce_display())
            ws.cell(row=row_idx, column=9, value="Ano" if t.vyplaceno else "Ne")
            ws.cell(row=row_idx, column=10, value=t.cislo_protiuctu or "")
            ws.cell(row=row_idx, column=11, value=t.nazev_protiuctu or "")
            ws.cell(row=row_idx, column=12, value=t.nazev_merchanta or "")
            ws.cell(row=row_idx, column=13, value=t.mesto or "")
            ws.cell(row=row_idx, column=14, value=t.get_status_display())
            ws.cell(row=row_idx, column=15, value=t.prijem_vydaj or "")
            ws.cell(row=row_idx, column=16, value=t.vlastni_nevlastni or "")
            ws.cell(row=row_idx, column=17, value="Ano" if t.dane else "Ne")
            ws.cell(row=row_idx, column=18, value=t.druh or "")
            ws.cell(row=row_idx, column=19, value=t.detail or "")
            ws.cell(row=row_idx, column=20, value=t.zodpovedna_osoba or "")
            ws.cell(row=row_idx, column=21, value=t.kmen or "")
            ws.cell(row=row_idx, column=22, value=float(t.mh_pct))
            ws.cell(row=row_idx, column=23, value=float(t.sk_pct))
            ws.cell(row=row_idx, column=24, value=float(t.xp_pct))
            ws.cell(row=row_idx, column=25, value=float(t.fr_pct))
            ws.cell(row=row_idx, column=26, value=t.projekt.name if t.projekt else "")
            ws.cell(row=row_idx, column=27, value=t.produkt.name if t.produkt else "")
            ws.cell(
                row=row_idx, column=28, value=t.podskupina.name if t.podskupina else ""
            )

        # Auto-width columns
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(headers[col_idx - 1]))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

        # Number format for amount column
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
            for cell in row:
                cell.number_format = "#,##0.00"

        # Freeze header row
        ws.freeze_panes = "A2"

        # Write to response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        from datetime import date as date_cls

        today = date_cls.today().strftime("%d_%m_%Y")
        filename = f"HeroWizzardTransakce{today}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=["post"], url_path="create-manual")
    def create_manual(self, request):
        """
        Manually create a single transaction.

        POST /api/v1/transactions/create-manual/
        Allows setting key bank columns (datum, castka, …) that are
        read-only on the standard update serializer.
        """
        serializer = ManualTransactionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        user_role = getattr(request.user, "role", "viewer")
        initial_status = (
            Transaction.Status.CEKA_NA_SCHVALENI
            if user_role in ("accountant", "viewer")
            else "upraveno"
        )
        transaction = serializer.save(
            created_by=request.user,
            updated_by=request.user,
            status=initial_status,
        )

        # Auto-set P/V from amount sign when not explicitly provided
        if not transaction.prijem_vydaj:
            transaction.prijem_vydaj = "P" if transaction.castka > 0 else "V"
            transaction.save(update_fields=["prijem_vydaj"])

        TransactionAuditLog.objects.create(
            transaction=transaction,
            user=request.user,
            action="Ruční vytvoření",
            details="",
        )

        return Response(
            TransactionDetailSerializer(transaction).data,
            status=status.HTTP_201_CREATED,
        )

    @action(
        detail=False,
        methods=["get"],
        url_path="export-backup",
        permission_classes=[IsAdminUser],
    )
    def export_backup(self, request):
        """
        Export full application state as JSON backup (admin only).
        Includes: lookups, transactions, category rules, audit logs, import batches.

        GET /api/v1/transactions/export-backup/
        """
        import json

        from django.http import HttpResponse

        # --- Transactions ---
        txn_qs = (
            Transaction.objects.filter(is_deleted=False)
            .select_related("projekt", "produkt", "podskupina")
            .order_by("datum", "created_at")
        )
        txn_records = []
        for t in txn_qs.iterator():
            txn_records.append(
                {
                    "id": str(t.id),
                    "datum": t.datum.isoformat() if t.datum else None,
                    "ucet": t.ucet,
                    "typ": t.typ,
                    "poznamka_zprava": t.poznamka_zprava,
                    "variabilni_symbol": t.variabilni_symbol,
                    "castka": str(t.castka),
                    "datum_zauctovani": (
                        t.datum_zauctovani.isoformat() if t.datum_zauctovani else None
                    ),
                    "cislo_protiuctu": t.cislo_protiuctu,
                    "nazev_protiuctu": t.nazev_protiuctu,
                    "typ_transakce": t.typ_transakce,
                    "konstantni_symbol": t.konstantni_symbol,
                    "specificky_symbol": t.specificky_symbol,
                    "puvodni_castka": (
                        str(t.puvodni_castka) if t.puvodni_castka else None
                    ),
                    "puvodni_mena": t.puvodni_mena,
                    "poplatky": str(t.poplatky) if t.poplatky else None,
                    "id_transakce": t.id_transakce,
                    "vlastni_poznamka": t.vlastni_poznamka,
                    "nazev_merchanta": t.nazev_merchanta,
                    "mesto": t.mesto,
                    "mena": t.mena,
                    "banka_protiuctu": t.banka_protiuctu,
                    "reference": t.reference,
                    "zdroj_transakce": t.zdroj_transakce,
                    "vyplaceno": t.vyplaceno,
                    "status": t.status,
                    "prijem_vydaj": t.prijem_vydaj,
                    "vlastni_nevlastni": t.vlastni_nevlastni,
                    "dane": t.dane,
                    "druh": t.druh,
                    "detail": t.detail,
                    "zodpovedna_osoba": t.zodpovedna_osoba,
                    "kmen": t.kmen,
                    "mh_pct": str(t.mh_pct),
                    "sk_pct": str(t.sk_pct),
                    "xp_pct": str(t.xp_pct),
                    "fr_pct": str(t.fr_pct),
                    "projekt": t.projekt.name if t.projekt else None,
                    "produkt": t.produkt.name if t.produkt else None,
                    "podskupina": t.podskupina.name if t.podskupina else None,
                    "is_active": t.is_active,
                    "import_batch_id": (
                        str(t.import_batch_id) if t.import_batch_id else None
                    ),
                    "created_at": t.created_at.isoformat() if t.created_at else None,
                }
            )

        # --- Category Rules ---
        rule_records = []
        for r in CategoryRule.objects.select_related(
            "set_projekt", "set_produkt", "set_podskupina"
        ).order_by("match_type", "priority"):
            rule_records.append(
                {
                    "id": str(r.id),
                    "name": r.name,
                    "description": r.description,
                    "match_type": r.match_type,
                    "match_mode": r.match_mode,
                    "match_value": r.match_value,
                    "case_sensitive": r.case_sensitive,
                    "priority": r.priority,
                    "set_prijem_vydaj": r.set_prijem_vydaj,
                    "set_vlastni_nevlastni": r.set_vlastni_nevlastni,
                    "set_dane": r.set_dane,
                    "set_druh": r.set_druh,
                    "set_detail": r.set_detail,
                    "set_kmen": r.set_kmen,
                    "set_mh_pct": (
                        str(r.set_mh_pct) if r.set_mh_pct is not None else None
                    ),
                    "set_sk_pct": (
                        str(r.set_sk_pct) if r.set_sk_pct is not None else None
                    ),
                    "set_xp_pct": (
                        str(r.set_xp_pct) if r.set_xp_pct is not None else None
                    ),
                    "set_fr_pct": (
                        str(r.set_fr_pct) if r.set_fr_pct is not None else None
                    ),
                    "set_projekt": r.set_projekt.name if r.set_projekt else None,
                    "set_produkt": r.set_produkt.name if r.set_produkt else None,
                    "set_podskupina": (
                        r.set_podskupina.name if r.set_podskupina else None
                    ),
                    "is_active": r.is_active,
                    "created_at": r.created_at.isoformat() if r.created_at else None,
                }
            )

        # --- Import Batches ---
        batch_records = []
        for b in ImportBatch.objects.order_by("created_at"):
            batch_records.append(
                {
                    "id": str(b.id),
                    "filename": b.filename,
                    "status": b.status,
                    "total_rows": b.total_rows,
                    "imported_count": b.imported_count,
                    "skipped_count": b.skipped_count,
                    "error_count": b.error_count,
                    "error_details": b.error_details,
                    "started_at": b.started_at.isoformat() if b.started_at else None,
                    "completed_at": (
                        b.completed_at.isoformat() if b.completed_at else None
                    ),
                    "created_at": b.created_at.isoformat() if b.created_at else None,
                }
            )

        # --- Audit Logs ---
        audit_records = []
        for a in TransactionAuditLog.objects.select_related("user").order_by(
            "created_at"
        ):
            audit_records.append(
                {
                    "id": str(a.id),
                    "transaction_id": str(a.transaction_id),
                    "user_email": a.user.email if a.user else None,
                    "action": a.action,
                    "details": a.details,
                    "created_at": a.created_at.isoformat() if a.created_at else None,
                }
            )

        # --- Module 2 data ---
        from apps.projects.models import BudgetCategory as M2BudgetCategory
        from apps.projects.models import \
            BudgetLineTemplate as M2BudgetLineTemplate
        from apps.projects.models import Client as M2Client
        from apps.projects.models import ClientCategory as M2ClientCategory
        from apps.projects.models import Deal as M2Deal
        from apps.projects.models import DealBudgetLine as M2DealBudgetLine
        from apps.projects.models import \
            DealPersonAssignment as M2DealPersonAssignment
        from apps.projects.models import DealStatus as M2DealStatus
        from apps.projects.models import DealType as M2DealType
        from apps.projects.models import Organization as M2Organization
        from apps.projects.models import Person as M2Person
        from apps.projects.models import PersonType as M2PersonType
        from apps.projects.models import ProductCatalog as M2ProductCatalog

        m2_organizations = [
            {
                "id": o.id,
                "name": o.name,
                "ico": o.ico,
                "dic": o.dic,
                "address": o.address,
                "email": o.email,
                "phone": o.phone,
                "bank_account": o.bank_account,
                "iban": o.iban,
                "swift": o.swift,
                "registration_text": o.registration_text,
                "sort_order": o.sort_order,
                "is_active": o.is_active,
            }
            for o in M2Organization.objects.order_by("sort_order")
        ]
        m2_person_types = [
            {
                "id": pt.id,
                "label": pt.label,
                "color": pt.color,
                "sort_order": pt.sort_order,
                "is_active": pt.is_active,
            }
            for pt in M2PersonType.objects.order_by("sort_order")
        ]
        m2_persons = [
            {
                "id": str(p.id),
                "name": p.name,
                "person_type_id": p.person_type_id or "",
                "role_description": p.role_description,
                "organization_id": p.organization_id or "",
                "monthly_salary": str(p.monthly_salary),
                "ico": p.ico,
                "email": p.email,
                "phone": p.phone,
                "notes": p.notes,
                "is_active": p.is_active,
                "user_email": p.user.email if p.user else None,
            }
            for p in M2Person.objects.select_related("user").order_by("name")
        ]
        m2_client_categories = [
            {
                "id": cc.id,
                "label": cc.label,
                "sort_order": cc.sort_order,
                "is_active": cc.is_active,
            }
            for cc in M2ClientCategory.objects.order_by("sort_order")
        ]
        m2_clients = [
            {
                "id": str(c.id),
                "name": c.name,
                "ico": c.ico,
                "dic": c.dic,
                "address": c.address,
                "contact_person": c.contact_person,
                "email": c.email,
                "phone": c.phone,
                "category_id": c.category_id or "",
                "notes": c.notes,
                "is_active": c.is_active,
            }
            for c in M2Client.objects.order_by("name")
        ]
        m2_deal_types = [
            {
                "id": dt.id,
                "label": dt.label,
                "color": dt.color,
                "sort_order": dt.sort_order,
                "is_active": dt.is_active,
            }
            for dt in M2DealType.objects.order_by("sort_order")
        ]
        m2_deal_statuses = [
            {
                "id": ds.id,
                "label": ds.label,
                "color": ds.color,
                "sort_order": ds.sort_order,
                "is_active": ds.is_active,
            }
            for ds in M2DealStatus.objects.order_by("sort_order")
        ]
        m2_budget_categories = [
            {
                "id": bc.id,
                "label": bc.label,
                "is_revenue": bc.is_revenue,
                "sort_order": bc.sort_order,
                "is_active": bc.is_active,
            }
            for bc in M2BudgetCategory.objects.order_by("is_revenue", "sort_order")
        ]

        m2_products = [
            {
                "id": str(p.id),
                "name": p.name,
                "deal_type_id": p.deal_type_id,
                "unit": p.unit,
                "default_price": str(p.default_price),
                "description": p.description,
                "sort_order": p.sort_order,
                "is_active": p.is_active,
            }
            for p in M2ProductCatalog.objects.order_by("deal_type", "sort_order")
        ]
        m2_budget_templates = [
            {
                "id": str(t.id),
                "deal_type_id": t.deal_type_id,
                "category_id": t.category_id,
                "label": t.label,
                "calculation_rule": t.calculation_rule,
                "default_amount": str(t.default_amount),
                "sort_order": t.sort_order,
                "is_active": t.is_active,
            }
            for t in M2BudgetLineTemplate.objects.order_by("deal_type", "sort_order")
        ]
        m2_deals = [
            {
                "id": str(d.id),
                "title": d.title,
                "deal_type_id": d.deal_type_id,
                "client_id": str(d.client_id),
                "status_id": d.status_id,
                "organization_id": d.organization_id,
                "product_id": str(d.product_id) if d.product_id else None,
                "projekt_tag_id": d.projekt_tag_id or None,
                "owner_email": d.owner.email if d.owner else None,
                "revenue": str(d.revenue),
                "cost": str(d.cost),
                "margin": str(d.margin),
                "quantity": d.quantity,
                "unit_price": str(d.unit_price),
                "organization_split": d.organization_split,
                "date_start": d.date_start.isoformat() if d.date_start else None,
                "date_end": d.date_end.isoformat() if d.date_end else None,
                "notes": d.notes,
                "is_active": d.is_active,
                "created_at": d.created_at.isoformat() if d.created_at else None,
            }
            for d in M2Deal.objects.select_related("owner").order_by("created_at")
        ]
        m2_deal_budget_lines = [
            {
                "id": str(bl.id),
                "deal_id": str(bl.deal_id),
                "category_id": bl.category_id,
                "label": bl.label,
                "calculation_rule": bl.calculation_rule,
                "budgeted": str(bl.budgeted),
                "actual": str(bl.actual),
                "sort_order": bl.sort_order,
                "is_active": bl.is_active,
            }
            for bl in M2DealBudgetLine.objects.order_by("deal", "sort_order")
        ]
        m2_deal_person_assignments = [
            {
                "id": str(a.id),
                "deal_id": str(a.deal_id),
                "person_id": str(a.person_id),
                "role": a.role,
                "is_active": a.is_active,
            }
            for a in M2DealPersonAssignment.objects.order_by("deal", "person")
        ]

        # --- Lookups (Projects, Products, Subgroups) ---
        project_records = [
            {
                "id": p.id,
                "name": p.name,
                "description": p.description,
                "sort_order": p.sort_order,
                "is_active": p.is_active,
            }
            for p in Project.objects.order_by("sort_order", "name")
        ]
        product_records = [
            {
                "id": p.id,
                "name": p.name,
                "category": p.category,
                "description": p.description,
                "sort_order": p.sort_order,
                "is_active": p.is_active,
            }
            for p in Product.objects.order_by("sort_order", "category", "name")
        ]
        subgroup_records = [
            {
                "id": s.id,
                "product_id": s.product_id,
                "name": s.name,
                "description": s.description,
                "sort_order": s.sort_order,
                "is_active": s.is_active,
            }
            for s in ProductSubgroup.objects.order_by("product", "sort_order", "name")
        ]
        cost_detail_records = [
            {
                "id": cd.id,
                "druh_type": cd.druh_type,
                "druh_value": cd.druh_value,
                "detail": cd.detail,
                "poznamka": cd.poznamka,
                "sort_order": cd.sort_order,
                "is_active": cd.is_active,
            }
            for cd in CostDetail.objects.order_by(
                "druh_type", "sort_order", "druh_value"
            )
        ]

        payload = json.dumps(
            {
                "version": 8,
                "projects": project_records,
                "products": product_records,
                "product_subgroups": subgroup_records,
                "cost_details": cost_detail_records,
                "transactions": txn_records,
                "category_rules": rule_records,
                "import_batches": batch_records,
                "audit_logs": audit_records,
                "m2_organizations": m2_organizations,
                "m2_person_types": m2_person_types,
                "m2_persons": m2_persons,
                "m2_client_categories": m2_client_categories,
                "m2_clients": m2_clients,
                "m2_deal_types": m2_deal_types,
                "m2_deal_statuses": m2_deal_statuses,
                "m2_budget_categories": m2_budget_categories,
                "m2_products": m2_products,
                "m2_budget_templates": m2_budget_templates,
                "m2_deals": m2_deals,
                "m2_deal_budget_lines": m2_deal_budget_lines,
                "m2_deal_person_assignments": m2_deal_person_assignments,
                "counts": {
                    "projects": len(project_records),
                    "products": len(product_records),
                    "product_subgroups": len(subgroup_records),
                    "cost_details": len(cost_detail_records),
                    "transactions": len(txn_records),
                    "category_rules": len(rule_records),
                    "import_batches": len(batch_records),
                    "audit_logs": len(audit_records),
                    "m2_organizations": len(m2_organizations),
                    "m2_person_types": len(m2_person_types),
                    "m2_persons": len(m2_persons),
                    "m2_client_categories": len(m2_client_categories),
                    "m2_clients": len(m2_clients),
                    "m2_deal_types": len(m2_deal_types),
                    "m2_deal_statuses": len(m2_deal_statuses),
                    "m2_budget_categories": len(m2_budget_categories),
                    "m2_products": len(m2_products),
                    "m2_budget_templates": len(m2_budget_templates),
                    "m2_deals": len(m2_deals),
                    "m2_deal_budget_lines": len(m2_deal_budget_lines),
                    "m2_deal_person_assignments": len(m2_deal_person_assignments),
                },
            },
            ensure_ascii=False,
            indent=2,
        )

        response = HttpResponse(payload, content_type="application/json; charset=utf-8")
        response["Content-Disposition"] = (
            'attachment; filename="herowizzard_backup.json"'
        )
        return response

    @action(
        detail=False,
        methods=["post"],
        url_path="import-backup",
        parser_classes=[MultiPartParser, FormParser],
        permission_classes=[IsAdminUser],
    )
    def import_backup(self, request):
        """
        Restore application state from a JSON backup file (admin only).
        DELETES all existing transactions, category rules, audit logs,
        and import batches, then recreates them from the backup.
        Users and roles are NOT affected.

        Supports v3 (transactions only), v5 (full backup), v6 (+ cost details), v7 (+ Module 2 Phase 1), and v8 (+ Module 2 Phase 2a: deals, products, templates) formats.

        POST /api/v1/transactions/import-backup/
        """
        import json

        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response(
                {"success": False, "error": "Nebyl nahrán žádný soubor."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            content = uploaded_file.read().decode("utf-8")
            data = json.loads(content)
        except (UnicodeDecodeError, json.JSONDecodeError) as e:
            return Response(
                {"success": False, "error": f"Neplatný JSON soubor: {e}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return self._restore_from_data(request, data)

    def _restore_from_data(self, request, data):
        """Shared restore logic used by both import_backup and restore_server_backup."""
        import uuid as _uuid
        from datetime import date

        from django.db import connection
        from django.db import transaction as db_transaction
        from django.utils.dateparse import parse_datetime

        backup_version = data.get("version", 3)
        txn_records = data.get("transactions", [])
        rule_records = data.get("category_rules", [])
        batch_records = data.get("import_batches", [])
        audit_records = data.get("audit_logs", [])
        # v6+ lookups
        project_records = data.get("projects", [])
        product_records = data.get("products", [])
        subgroup_records = data.get("product_subgroups", [])
        cost_detail_records = data.get("cost_details", [])
        # v7+ Module 2
        m2_organizations = data.get("m2_organizations", [])
        m2_person_types = data.get("m2_person_types", [])
        m2_persons = data.get("m2_persons", [])
        m2_client_categories = data.get("m2_client_categories", [])
        m2_clients = data.get("m2_clients", [])
        m2_deal_types = data.get("m2_deal_types", [])
        m2_deal_statuses = data.get("m2_deal_statuses", [])
        m2_budget_categories = data.get("m2_budget_categories", [])
        # v8+ Module 2 Phase 2a
        m2_products = data.get("m2_products", [])
        m2_budget_templates = data.get("m2_budget_templates", [])
        m2_deals = data.get("m2_deals", [])
        m2_deal_budget_lines = data.get("m2_deal_budget_lines", [])
        m2_deal_person_assignments = data.get("m2_deal_person_assignments", [])

        if not txn_records and not rule_records:
            return Response(
                {"success": False, "error": "Záloha neobsahuje žádná data."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Resolve user by email for audit logs
        from apps.core.models import User

        user_map = {u.email: u for u in User.objects.all()}

        errors = []
        counts = {
            "transactions_deleted": 0,
            "transactions_imported": 0,
            "rules_deleted": 0,
            "rules_imported": 0,
            "batches_deleted": 0,
            "batches_imported": 0,
            "audit_logs_deleted": 0,
            "audit_logs_imported": 0,
            "projects_imported": 0,
            "products_imported": 0,
            "subgroups_imported": 0,
            "cost_details_imported": 0,
        }

        try:
            with db_transaction.atomic():
                # ---- PHASE 1: DELETE existing data ----
                counts["audit_logs_deleted"] = TransactionAuditLog.objects.count()
                counts["transactions_deleted"] = Transaction.objects.count()
                counts["batches_deleted"] = ImportBatch.objects.count()
                counts["rules_deleted"] = CategoryRule.objects.count()
                # Use TRUNCATE CASCADE to handle FK constraints at DB level
                from django.db import connection as db_conn

                has_lookups = bool(
                    project_records or product_records or subgroup_records
                )
                has_cost_details = bool(cost_detail_records)
                has_m2 = bool(
                    m2_organizations
                    or m2_person_types
                    or m2_persons
                    or m2_client_categories
                    or m2_clients
                    or m2_deal_types
                    or m2_deal_statuses
                    or m2_budget_categories
                    or m2_products
                    or m2_budget_templates
                    or m2_deals
                    or m2_deal_budget_lines
                    or m2_deal_person_assignments
                )
                with db_conn.cursor() as cursor:
                    tables = (
                        "transactions_audit_log, "
                        "transactions_transaction, "
                        "transactions_import_batch, "
                        "transactions_category_rule"
                    )
                    if has_lookups:
                        tables += (
                            ", transactions_product_subgroup"
                            ", transactions_product"
                            ", transactions_project"
                        )
                    if has_cost_details:
                        tables += ", transactions_cost_detail"
                    if has_m2:
                        tables += (
                            ", projects_deal_person_assignment"
                            ", projects_deal_budget_line"
                            ", projects_deal"
                            ", projects_budget_line_template"
                            ", projects_product_catalog"
                            ", projects_person"
                            ", projects_client"
                            ", projects_organization"
                            ", projects_person_type"
                            ", projects_client_category"
                            ", projects_deal_type"
                            ", projects_deal_status"
                            ", projects_budget_category"
                        )
                    cursor.execute(f"TRUNCATE TABLE {tables} CASCADE")

                # ---- PHASE 1b: Restore lookups (v6+) ----
                if project_records:
                    for rec in project_records:
                        try:
                            Project.objects.create(
                                id=rec["id"],
                                name=rec["name"],
                                description=rec.get("description", ""),
                                sort_order=rec.get("sort_order", 0),
                                is_active=rec.get("is_active", True),
                            )
                            counts["projects_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "project",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )

                if product_records:
                    for rec in product_records:
                        try:
                            Product.objects.create(
                                id=rec["id"],
                                name=rec["name"],
                                category=rec.get("category", "SKOLY"),
                                description=rec.get("description", ""),
                                sort_order=rec.get("sort_order", 0),
                                is_active=rec.get("is_active", True),
                            )
                            counts["products_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "product",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )

                if subgroup_records:
                    for rec in subgroup_records:
                        try:
                            ProductSubgroup.objects.create(
                                id=rec["id"],
                                product_id=rec["product_id"],
                                name=rec["name"],
                                description=rec.get("description", ""),
                                sort_order=rec.get("sort_order", 0),
                                is_active=rec.get("is_active", True),
                            )
                            counts["subgroups_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "subgroup",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )

                if cost_detail_records:
                    for rec in cost_detail_records:
                        try:
                            CostDetail.objects.create(
                                id=rec["id"],
                                druh_type=rec["druh_type"],
                                druh_value=rec["druh_value"],
                                detail=rec.get("detail", ""),
                                poznamka=rec.get("poznamka", ""),
                                sort_order=rec.get("sort_order", 0),
                                is_active=rec.get("is_active", True),
                            )
                            counts["cost_details_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "cost_detail",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )

                # ---- PHASE 1c: Restore Module 2 data (v7+) ----
                if has_m2:
                    from apps.projects.models import \
                        BudgetCategory as M2BudgetCategory
                    from apps.projects.models import \
                        BudgetLineTemplate as M2BudgetLineTemplate
                    from apps.projects.models import Client as M2Client
                    from apps.projects.models import \
                        ClientCategory as M2ClientCategory
                    from apps.projects.models import Deal as M2Deal
                    from apps.projects.models import \
                        DealBudgetLine as M2DealBudgetLine
                    from apps.projects.models import \
                        DealPersonAssignment as M2DealPersonAssignment
                    from apps.projects.models import DealStatus as M2DealStatus
                    from apps.projects.models import DealType as M2DealType
                    from apps.projects.models import \
                        Organization as M2Organization
                    from apps.projects.models import Person as M2Person
                    from apps.projects.models import PersonType as M2PersonType
                    from apps.projects.models import \
                        ProductCatalog as M2ProductCatalog

                    # Lookups first (FKs depend on them)
                    for rec in m2_organizations:
                        try:
                            M2Organization.objects.create(**rec)
                            counts.setdefault("m2_organizations_imported", 0)
                            counts["m2_organizations_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "m2_organization",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )
                    for rec in m2_person_types:
                        try:
                            M2PersonType.objects.create(**rec)
                            counts.setdefault("m2_person_types_imported", 0)
                            counts["m2_person_types_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "m2_person_type",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )
                    for rec in m2_client_categories:
                        try:
                            M2ClientCategory.objects.create(**rec)
                            counts.setdefault("m2_client_categories_imported", 0)
                            counts["m2_client_categories_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "m2_client_category",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )
                    for rec in m2_deal_types:
                        try:
                            M2DealType.objects.create(**rec)
                            counts.setdefault("m2_deal_types_imported", 0)
                            counts["m2_deal_types_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "m2_deal_type",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )
                    for rec in m2_deal_statuses:
                        try:
                            M2DealStatus.objects.create(**rec)
                            counts.setdefault("m2_deal_statuses_imported", 0)
                            counts["m2_deal_statuses_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "m2_deal_status",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )
                    for rec in m2_budget_categories:
                        try:
                            M2BudgetCategory.objects.create(**rec)
                            counts.setdefault("m2_budget_categories_imported", 0)
                            counts["m2_budget_categories_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "m2_budget_category",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )
                    # Entities (depend on lookups)
                    for rec in m2_persons:
                        try:
                            user_email = rec.pop("user_email", None)
                            user = user_map.get(user_email) if user_email else None
                            M2Person.objects.create(user=user, **rec)
                            counts.setdefault("m2_persons_imported", 0)
                            counts["m2_persons_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "m2_person",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )
                    for rec in m2_clients:
                        try:
                            M2Client.objects.create(**rec)
                            counts.setdefault("m2_clients_imported", 0)
                            counts["m2_clients_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "m2_client",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )

                    # Phase 2a: ProductCatalog (depends on DealType)
                    for rec in m2_products:
                        try:
                            M2ProductCatalog.objects.create(
                                id=rec["id"],
                                name=rec["name"],
                                deal_type_id=rec["deal_type_id"],
                                unit=rec.get("unit", ""),
                                default_price=Decimal(rec.get("default_price", "0")),
                                description=rec.get("description", ""),
                                sort_order=rec.get("sort_order", 0),
                                is_active=rec.get("is_active", True),
                            )
                            counts.setdefault("m2_products_imported", 0)
                            counts["m2_products_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "m2_product",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )
                    # BudgetLineTemplate (depends on DealType, BudgetCategory)
                    for rec in m2_budget_templates:
                        try:
                            M2BudgetLineTemplate.objects.create(
                                id=rec["id"],
                                deal_type_id=rec["deal_type_id"],
                                category_id=rec["category_id"],
                                label=rec["label"],
                                calculation_rule=rec.get("calculation_rule", "fixed"),
                                default_amount=Decimal(rec.get("default_amount", "0")),
                                sort_order=rec.get("sort_order", 0),
                                is_active=rec.get("is_active", True),
                            )
                            counts.setdefault("m2_budget_templates_imported", 0)
                            counts["m2_budget_templates_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "m2_budget_template",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )
                    # Deal (depends on DealType, Client, DealStatus, Organization, ProductCatalog)
                    for rec in m2_deals:
                        try:
                            owner_email = rec.pop("owner_email", None)
                            owner = user_map.get(owner_email) if owner_email else None
                            M2Deal.objects.create(
                                id=rec["id"],
                                title=rec["title"],
                                deal_type_id=rec["deal_type_id"],
                                client_id=rec["client_id"],
                                status_id=rec["status_id"],
                                organization_id=rec["organization_id"],
                                product_id=rec.get("product_id") or None,
                                projekt_tag_id=rec.get("projekt_tag_id") or None,
                                owner=owner,
                                revenue=Decimal(rec.get("revenue", "0")),
                                cost=Decimal(rec.get("cost", "0")),
                                margin=Decimal(rec.get("margin", "0")),
                                quantity=rec.get("quantity", 1),
                                unit_price=Decimal(rec.get("unit_price", "0")),
                                organization_split=rec.get("organization_split", {}),
                                date_start=rec.get("date_start") or None,
                                date_end=rec.get("date_end") or None,
                                notes=rec.get("notes", ""),
                                is_active=rec.get("is_active", True),
                            )
                            counts.setdefault("m2_deals_imported", 0)
                            counts["m2_deals_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "m2_deal",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )
                    # DealBudgetLine (depends on Deal, BudgetCategory)
                    for rec in m2_deal_budget_lines:
                        try:
                            M2DealBudgetLine.objects.create(
                                id=rec["id"],
                                deal_id=rec["deal_id"],
                                category_id=rec["category_id"],
                                label=rec["label"],
                                calculation_rule=rec.get("calculation_rule", "fixed"),
                                budgeted=Decimal(rec.get("budgeted", "0")),
                                actual=Decimal(rec.get("actual", "0")),
                                sort_order=rec.get("sort_order", 0),
                                is_active=rec.get("is_active", True),
                            )
                            counts.setdefault("m2_deal_budget_lines_imported", 0)
                            counts["m2_deal_budget_lines_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "m2_deal_budget_line",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )
                    # DealPersonAssignment (depends on Deal, Person)
                    for rec in m2_deal_person_assignments:
                        try:
                            M2DealPersonAssignment.objects.create(
                                id=rec["id"],
                                deal_id=rec["deal_id"],
                                person_id=rec["person_id"],
                                role=rec.get("role", ""),
                                is_active=rec.get("is_active", True),
                            )
                            counts.setdefault("m2_deal_person_assignments_imported", 0)
                            counts["m2_deal_person_assignments_imported"] += 1
                        except Exception as e:
                            errors.append(
                                {
                                    "type": "m2_deal_person_assignment",
                                    "id": rec.get("id"),
                                    "error": str(e),
                                }
                            )

                # Build lookup maps after restore
                project_map = {p.name: p for p in Project.objects.all()}
                product_map = {p.name: p for p in Product.objects.all()}
                subgroup_map = {s.name: s for s in ProductSubgroup.objects.all()}

                # ---- PHASE 2: Restore import batches first (transactions reference them) ----
                for rec in batch_records:
                    try:
                        batch = ImportBatch(
                            id=_uuid.UUID(rec["id"]),
                            filename=rec.get("filename", ""),
                            status=rec.get("status", "completed"),
                            total_rows=rec.get("total_rows", 0),
                            imported_count=rec.get("imported_count", 0),
                            skipped_count=rec.get("skipped_count", 0),
                            error_count=rec.get("error_count", 0),
                            error_details=rec.get("error_details", []),
                        )
                        batch.save()
                        # Update timestamps via raw SQL (auto_now_add prevents direct set)
                        if rec.get("created_at"):
                            ImportBatch.objects.filter(id=batch.id).update(
                                created_at=parse_datetime(rec["created_at"]),
                                started_at=(
                                    parse_datetime(rec["started_at"])
                                    if rec.get("started_at")
                                    else None
                                ),
                                completed_at=(
                                    parse_datetime(rec["completed_at"])
                                    if rec.get("completed_at")
                                    else None
                                ),
                            )
                        counts["batches_imported"] += 1
                    except Exception as e:
                        errors.append(
                            {"type": "batch", "id": rec.get("id"), "error": str(e)}
                        )

                # ---- PHASE 3: Restore category rules ----
                for rec in rule_records:
                    try:
                        rule = CategoryRule(
                            id=_uuid.UUID(rec["id"]),
                            name=rec.get("name", ""),
                            description=rec.get("description", ""),
                            match_type=rec.get("match_type", ""),
                            match_mode=rec.get("match_mode", "exact"),
                            match_value=rec.get("match_value", ""),
                            case_sensitive=rec.get("case_sensitive", False),
                            priority=rec.get("priority", 100),
                            set_prijem_vydaj=rec.get("set_prijem_vydaj", ""),
                            set_vlastni_nevlastni=rec.get("set_vlastni_nevlastni", ""),
                            set_dane=rec.get("set_dane"),
                            set_druh=rec.get("set_druh", ""),
                            set_detail=rec.get("set_detail", ""),
                            set_kmen=rec.get("set_kmen", ""),
                            set_mh_pct=(
                                Decimal(rec["set_mh_pct"])
                                if rec.get("set_mh_pct") is not None
                                else None
                            ),
                            set_sk_pct=(
                                Decimal(rec["set_sk_pct"])
                                if rec.get("set_sk_pct") is not None
                                else None
                            ),
                            set_xp_pct=(
                                Decimal(rec["set_xp_pct"])
                                if rec.get("set_xp_pct") is not None
                                else None
                            ),
                            set_fr_pct=(
                                Decimal(rec["set_fr_pct"])
                                if rec.get("set_fr_pct") is not None
                                else None
                            ),
                            is_active=rec.get("is_active", True),
                            created_by=request.user,
                        )
                        if rec.get("set_projekt"):
                            rule.set_projekt = project_map.get(rec["set_projekt"])
                        if rec.get("set_produkt"):
                            rule.set_produkt = product_map.get(rec["set_produkt"])
                        if rec.get("set_podskupina"):
                            rule.set_podskupina = subgroup_map.get(
                                rec["set_podskupina"]
                            )
                        rule.save()
                        # Restore original created_at
                        if rec.get("created_at"):
                            CategoryRule.objects.filter(id=rule.id).update(
                                created_at=parse_datetime(rec["created_at"])
                            )
                        counts["rules_imported"] += 1
                    except Exception as e:
                        errors.append(
                            {"type": "rule", "id": rec.get("id"), "error": str(e)}
                        )

                # ---- PHASE 4: Restore transactions ----
                for idx, rec in enumerate(txn_records, 1):
                    try:
                        txn = Transaction()
                        rec_id = rec.get("id", "")
                        if rec_id:
                            txn.id = _uuid.UUID(rec_id)
                        txn.datum = (
                            date.fromisoformat(rec["datum"])
                            if rec.get("datum")
                            else None
                        )
                        txn.ucet = rec.get("ucet", "")
                        txn.typ = rec.get("typ", "")
                        txn.poznamka_zprava = rec.get("poznamka_zprava", "")
                        txn.variabilni_symbol = rec.get("variabilni_symbol", "")
                        txn.castka = (
                            Decimal(rec["castka"])
                            if rec.get("castka")
                            else Decimal("0")
                        )
                        txn.datum_zauctovani = (
                            date.fromisoformat(rec["datum_zauctovani"])
                            if rec.get("datum_zauctovani")
                            else None
                        )
                        txn.cislo_protiuctu = rec.get("cislo_protiuctu", "")
                        txn.nazev_protiuctu = rec.get("nazev_protiuctu", "")
                        txn.typ_transakce = rec.get("typ_transakce", "")
                        txn.konstantni_symbol = rec.get("konstantni_symbol", "")
                        txn.specificky_symbol = rec.get("specificky_symbol", "")
                        txn.puvodni_castka = (
                            Decimal(rec["puvodni_castka"])
                            if rec.get("puvodni_castka")
                            else None
                        )
                        txn.puvodni_mena = rec.get("puvodni_mena", "")
                        txn.poplatky = (
                            Decimal(rec["poplatky"]) if rec.get("poplatky") else None
                        )
                        txn.id_transakce = rec.get("id_transakce", "")
                        txn.vlastni_poznamka = rec.get("vlastni_poznamka", "")
                        txn.nazev_merchanta = rec.get("nazev_merchanta", "")
                        txn.mesto = rec.get("mesto", "")
                        txn.mena = rec.get("mena", "CZK")
                        txn.banka_protiuctu = rec.get("banka_protiuctu", "")
                        txn.reference = rec.get("reference", "")
                        txn.zdroj_transakce = rec.get("zdroj_transakce", "ucet")
                        txn.vyplaceno = rec.get("vyplaceno", False)
                        txn.status = rec.get("status", "")
                        txn.prijem_vydaj = rec.get("prijem_vydaj", "")
                        txn.vlastni_nevlastni = rec.get("vlastni_nevlastni", "")
                        txn.dane = rec.get("dane", False)
                        txn.druh = rec.get("druh", "")
                        txn.detail = rec.get("detail", "")
                        txn.zodpovedna_osoba = rec.get("zodpovedna_osoba", "")
                        txn.kmen = rec.get("kmen", "")
                        txn.mh_pct = Decimal(rec.get("mh_pct", "0"))
                        txn.sk_pct = Decimal(rec.get("sk_pct", "0"))
                        txn.xp_pct = Decimal(rec.get("xp_pct", "0"))
                        txn.fr_pct = Decimal(rec.get("fr_pct", "0"))
                        txn.is_active = rec.get("is_active", True)
                        txn.is_deleted = False
                        if rec.get("projekt"):
                            txn.projekt = project_map.get(rec["projekt"])
                        if rec.get("produkt"):
                            txn.produkt = product_map.get(rec["produkt"])
                        if rec.get("podskupina"):
                            txn.podskupina = subgroup_map.get(rec["podskupina"])
                        txn.created_by = request.user
                        if rec.get("import_batch_id"):
                            txn.import_batch_id = _uuid.UUID(rec["import_batch_id"])
                        txn.save()
                        # Restore original created_at
                        if rec.get("created_at"):
                            Transaction.objects.filter(id=txn.id).update(
                                created_at=parse_datetime(rec["created_at"])
                            )
                        counts["transactions_imported"] += 1
                    except Exception as e:
                        errors.append(
                            {
                                "type": "transaction",
                                "row": idx,
                                "id": rec.get("id", ""),
                                "error": str(e),
                            }
                        )
                        if len(errors) > 50:
                            break

                # ---- PHASE 5: Restore audit logs ----
                # Build set of imported transaction IDs to skip orphaned audit logs
                # (e.g. logs for soft-deleted transactions not included in backup)
                imported_txn_ids = {
                    str(uid)
                    for uid in Transaction.objects.values_list(
                        "id", flat=True
                    ).iterator()
                }
                skipped_audit = 0
                for rec in audit_records:
                    try:
                        txn_id = rec.get("transaction_id", "")
                        if txn_id not in imported_txn_ids:
                            skipped_audit += 1
                            continue
                        log = TransactionAuditLog(
                            id=_uuid.UUID(rec["id"]),
                            transaction_id=_uuid.UUID(txn_id),
                            user=user_map.get(rec.get("user_email")),
                            action=rec.get("action", ""),
                            details=rec.get("details", ""),
                        )
                        log.save()
                        if rec.get("created_at"):
                            TransactionAuditLog.objects.filter(id=log.id).update(
                                created_at=parse_datetime(rec["created_at"])
                            )
                        counts["audit_logs_imported"] += 1
                    except Exception as e:
                        errors.append(
                            {"type": "audit_log", "id": rec.get("id"), "error": str(e)}
                        )
                if skipped_audit:
                    counts["audit_logs_skipped"] = skipped_audit

        except Exception as e:
            return Response(
                {"success": False, "error": f"Obnova selhala (rollback): {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response(
            {
                "success": True,
                "backup_version": backup_version,
                "counts": counts,
                "errors": len(errors),
                "error_details": errors[:10],
            },
            status=status.HTTP_201_CREATED,
        )

    @action(
        detail=False,
        methods=["get"],
        url_path="server-backups",
        permission_classes=[IsAdminUser],
    )
    def server_backups(self, request):
        """
        List available automatic backup files on the server (admin only).

        GET /api/v1/transactions/server-backups/
        """
        import os
        from datetime import datetime
        from pathlib import Path

        from django.conf import settings

        backup_dir = os.path.join(settings.BASE_DIR, "backups")
        if not os.path.isdir(backup_dir):
            return Response({"backups": []})

        backups = []
        for f in sorted(
            Path(backup_dir).glob("backup_*.json"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        ):
            stat = f.stat()
            backups.append(
                {
                    "filename": f.name,
                    "size_kb": round(stat.st_size / 1024, 1),
                    "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                }
            )

        return Response({"backups": backups})

    @action(
        detail=False,
        methods=["post"],
        url_path="restore-server-backup",
        permission_classes=[IsAdminUser],
    )
    def restore_server_backup(self, request):
        """
        Restore from a server-side automatic backup file (admin only).
        Expects JSON body: {"filename": "backup_2026-03-08_02-00.json"}

        POST /api/v1/transactions/restore-server-backup/
        """
        import json
        import os

        from django.conf import settings

        filename = request.data.get("filename", "")
        if not filename or ".." in filename or "/" in filename or "\\" in filename:
            return Response(
                {"error": "Neplatný název souboru."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        backup_dir = os.path.join(settings.BASE_DIR, "backups")
        filepath = os.path.join(backup_dir, filename)

        if not os.path.isfile(filepath):
            return Response(
                {"error": f"Soubor '{filename}' nebyl nalezen."},
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, UnicodeDecodeError) as e:
            return Response(
                {"error": f"Soubor nelze přečíst: {e}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return self._restore_from_data(request, data)

    @action(
        detail=False,
        methods=["post"],
        url_path="wipe-all",
        permission_classes=[IsAdminUser],
    )
    def wipe_all(self, request):
        """
        Soft-delete ALL non-deleted transactions (admin only).
        Sets is_deleted=True instead of actual deletion.

        POST /api/v1/transactions/wipe-all/
        """
        count = Transaction.objects.filter(is_deleted=False).update(is_deleted=True)

        return Response(
            {
                "success": True,
                "wiped_count": count,
                "message": f"Soft-deleted {count} transactions.",
            }
        )


# =============================================================================
# CATEGORY RULE VIEWSET
# =============================================================================


class CategoryRuleViewSet(viewsets.ModelViewSet):
    """
    API endpoint for CategoryRule management.

    list: Get all rules ordered by match_type and priority
    create: Create new rule
    update: Update rule
    delete: Delete or deactivate rule
    test: Test a rule against sample transactions
    apply: Apply rules to uncategorized transactions
    """

    queryset = CategoryRule.objects.select_related(
        "set_projekt", "set_produkt", "set_podskupina", "created_by"
    ).order_by("match_type", "priority", "name")
    serializer_class = CategoryRuleSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["match_type", "match_mode", "is_active"]
    search_fields = ["name", "description", "match_value"]

    def perform_create(self, serializer):
        """Set created_by to current user."""
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        """
        Export category rules as Excel (.xlsx).

        GET /api/v1/category-rules/export-excel/
        """
        from datetime import date as date_cls
        from io import BytesIO

        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        qs = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Pravidla kategorií"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="2563EB", end_color="2563EB", fill_type="solid"
        )
        header_align = Alignment(horizontal="center", vertical="center")

        headers = [
            "Název",
            "Popis",
            "Typ shody",
            "Režim shody",
            "Hodnota shody",
            "Priorita",
            "Aktivní",
            "Nastaví P/V",
            "Nastaví V/N",
            "Nastaví Daně",
            "Nastaví Druh",
            "Nastaví Detail",
            "Nastaví KMEN",
            "MH%",
            "ŠK%",
            "XP%",
            "FR%",
            "Nastaví Projekt",
            "Nastaví Produkt",
            "Nastaví Podskupinu",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row_idx, r in enumerate(qs.iterator(), 2):
            ws.cell(row=row_idx, column=1, value=r.name)
            ws.cell(row=row_idx, column=2, value=r.description or "")
            ws.cell(row=row_idx, column=3, value=r.get_match_type_display())
            ws.cell(row=row_idx, column=4, value=r.get_match_mode_display())
            ws.cell(row=row_idx, column=5, value=r.match_value)
            ws.cell(row=row_idx, column=6, value=r.priority)
            ws.cell(row=row_idx, column=7, value="Ano" if r.is_active else "Ne")
            ws.cell(row=row_idx, column=8, value=r.set_prijem_vydaj or "")
            ws.cell(row=row_idx, column=9, value=r.set_vlastni_nevlastni or "")
            ws.cell(
                row=row_idx,
                column=10,
                value="Ano" if r.set_dane else ("Ne" if r.set_dane is False else ""),
            )
            ws.cell(row=row_idx, column=11, value=r.set_druh or "")
            ws.cell(row=row_idx, column=12, value=r.set_detail or "")
            ws.cell(row=row_idx, column=13, value=r.set_kmen or "")
            ws.cell(
                row=row_idx,
                column=14,
                value=float(r.set_mh_pct) if r.set_mh_pct is not None else "",
            )
            ws.cell(
                row=row_idx,
                column=15,
                value=float(r.set_sk_pct) if r.set_sk_pct is not None else "",
            )
            ws.cell(
                row=row_idx,
                column=16,
                value=float(r.set_xp_pct) if r.set_xp_pct is not None else "",
            )
            ws.cell(
                row=row_idx,
                column=17,
                value=float(r.set_fr_pct) if r.set_fr_pct is not None else "",
            )
            ws.cell(
                row=row_idx,
                column=18,
                value=r.set_projekt.name if r.set_projekt else "",
            )
            ws.cell(
                row=row_idx,
                column=19,
                value=r.set_produkt.name if r.set_produkt else "",
            )
            ws.cell(
                row=row_idx,
                column=20,
                value=r.set_podskupina.name if r.set_podskupina else "",
            )

        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(headers[col_idx - 1]))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

        ws.freeze_panes = "A2"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        today = date_cls.today().strftime("%d_%m_%Y")
        filename = f"HeroWizzardPravidlaKategorii{today}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=["post"])
    def test(self, request, pk=None):
        """
        Test a rule against transactions.

        POST /api/v1/category-rules/{id}/test/
        Returns count of transactions that would match.
        """
        rule = self.get_object()
        importer = TransactionImporter(user=request.user)
        importer._load_caches()

        # Find matching transactions
        qs = Transaction.objects.all()
        match_count = 0
        sample_matches = []

        for txn in qs[:1000]:  # Limit to first 1000 for performance
            search_value = ""
            if rule.match_type == CategoryRule.MatchType.PROTIUCET:
                search_value = txn.cislo_protiuctu or ""
            elif rule.match_type == CategoryRule.MatchType.MERCHANT:
                search_value = txn.nazev_merchanta or ""
            elif rule.match_type == CategoryRule.MatchType.VS:
                search_value = txn.variabilni_symbol or ""
            elif rule.match_type == CategoryRule.MatchType.TYP:
                search_value = txn.typ or ""
            elif rule.match_type == CategoryRule.MatchType.MESTO:
                search_value = txn.mesto or ""
            elif rule.match_type == CategoryRule.MatchType.KEYWORD:
                search_value = " ".join(
                    filter(
                        None,
                        [
                            txn.poznamka_zprava,
                            txn.vlastni_poznamka,
                            txn.nazev_protiuctu,
                        ],
                    )
                )

            if search_value and importer._rule_matches(rule, search_value):
                match_count += 1
                if len(sample_matches) < 5:
                    sample_matches.append(
                        {
                            "id": str(txn.id),
                            "datum": txn.datum,
                            "castka": txn.castka,
                            "matched_text": search_value[:100],
                        }
                    )

        return Response(
            {
                "rule_id": str(rule.id),
                "rule_name": rule.name,
                "match_count": match_count,
                "sample_matches": sample_matches,
            }
        )

    @action(detail=False, methods=["post"])
    def apply_to_uncategorized(self, request):
        """
        Apply all active rules to uncategorized transactions.

        POST /api/v1/category-rules/apply_to_uncategorized/
        """
        importer = TransactionImporter(user=request.user)
        importer._load_caches()

        # Find uncategorized transactions
        uncategorized = list(
            Transaction.objects.filter(Q(prijem_vydaj="") | Q(druh=""))
        )

        updated_count = 0
        tracked_fields = [
            "prijem_vydaj",
            "vlastni_nevlastni",
            "dane",
            "druh",
            "detail",
            "kmen",
            "mh_pct",
            "sk_pct",
            "xp_pct",
            "fr_pct",
            "projekt_id",
            "produkt_id",
            "podskupina_id",
        ]
        for txn in uncategorized:
            original = {f: getattr(txn, f) for f in tracked_fields}
            txn = importer.apply_autodetection_rules(txn)
            current = {f: getattr(txn, f) for f in tracked_fields}

            if current != original:
                txn.save()
                updated_count += 1

        return Response(
            {
                "success": True,
                "processed_count": len(uncategorized),
                "updated_count": updated_count,
            }
        )


# =============================================================================
# IMPORT VIEWSET
# =============================================================================


class ImportBatchViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for ImportBatch management.

    list: Get all import batches
    retrieve: Get single batch with error details
    upload: Upload and process new CSV file
    """

    queryset = ImportBatch.objects.select_related("created_by").order_by("-created_at")
    serializer_class = ImportBatchSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["status"]
    ordering_fields = ["created_at", "total_rows", "imported_count"]

    @action(
        detail=False, methods=["post"], parser_classes=[MultiPartParser, FormParser]
    )
    def upload(self, request):
        """
        Upload and process a CSV file.

        POST /api/v1/imports/upload/
        Content-Type: multipart/form-data
        file: <csv_file>
        """
        serializer = CSVUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        uploaded_file = serializer.validated_data["file"]

        # Process the import
        importer = TransactionImporter(user=request.user)

        try:
            summary = importer.import_csv(
                file_stream=uploaded_file,
                filename=uploaded_file.name,
            )

            return Response(
                {
                    "success": True,
                    "batch_id": str(summary.batch_id),
                    "total_rows": summary.total_rows,
                    "imported": summary.imported,
                    "skipped": summary.skipped,
                    "errors": summary.errors,
                    "duration_seconds": summary.duration_seconds,
                    "error_details": summary.error_details[:10],  # First 10 errors
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "error": str(e),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["get"])
    def transactions(self, request, pk=None):
        """
        Get transactions from a specific import batch.

        GET /api/v1/imports/{id}/transactions/
        """
        batch = self.get_object()
        transactions = Transaction.objects.filter(import_batch_id=batch.id).order_by(
            "datum"
        )

        serializer = TransactionListSerializer(transactions, many=True)
        return Response(serializer.data)

    @action(
        detail=False,
        methods=["post"],
        url_path="upload-idoklad",
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_idoklad(self, request):
        """
        Upload and process an iDoklad invoice CSV file.

        POST /api/v1/imports/upload-idoklad/
        Content-Type: multipart/form-data
        file: <csv_file>
        """
        serializer = CSVUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        uploaded_file = serializer.validated_data["file"]

        importer = IDokladImporter(user=request.user)

        try:
            summary = importer.import_csv(
                file_stream=uploaded_file,
                filename=uploaded_file.name,
            )

            return Response(
                {
                    "success": True,
                    "batch_id": str(summary.batch_id),
                    "total_rows": summary.total_rows,
                    "imported": summary.imported,
                    "skipped": summary.skipped,
                    "errors": summary.errors,
                    "duration_seconds": summary.duration_seconds,
                    "error_details": summary.error_details[:10],
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "error": str(e),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
